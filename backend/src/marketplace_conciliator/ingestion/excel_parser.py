"""ExcelParser -- SourceParser adapter for XLSX/XLSM files (T-3.3 / ADR-004).

Key constraints from the spec:
  - Read-only mode: openpyxl ``read_only=True`` -- VBA/macros NEVER evaluated (EB-06).
  - All cells treated as strings: cell-by-cell conversion via openpyxl so that
    Excel error sentinels (#N/A, #REF!, etc.) survive as literal strings (ADR-004).
  - ``keep_default_na=False`` / ``na_filter=False`` are NOT sufficient on their
    own because openpyxl converts error-value cells to None before pandas sees
    them; we must stringify at the openpyxl layer.
  - Sheet listing without full cell parse.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

import openpyxl
import pandas as pd
from openpyxl.cell.cell import MergedCell

from marketplace_conciliator.ingestion.source_parser import (
    ParsedSource,
    UnsupportedFormatError,
)

if TYPE_CHECKING:
    from pathlib import Path

_SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})

# Canonical SKU column name patterns (case-insensitive fragments)
_SKU_HINTS: tuple[str, ...] = ("sku", "asin", "item_id", "product_id", "cod", "codigo")

# Placeholder for columns that have no header cell value
_UNNAMED_PLACEHOLDER = "(sin nombre)"


def _guess_sku_column(columns: list[str]) -> str | None:
    for col in columns:
        lower = col.lower()
        if any(hint in lower for hint in _SKU_HINTS):
            return col
    return None


def _cell_value_to_str(value: Any) -> str:  # noqa: ANN401
    """Convert an openpyxl cell value to a string, preserving error sentinels.

    openpyxl exposes cached error strings (e.g. '#N/A') as the raw string when
    ``data_only=True`` is used together with a workbook that has cached values.
    When there is no cached value the cell returns None, which maps to ''.
    """
    if value is None:
        return ""
    return str(value)


def _read_sheet_as_dataframe(wb: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
    """Read a single worksheet cell-by-cell and return a DataFrame of strings.

    This approach bypasses pandas' NaN-inference pipeline entirely -- every cell
    is converted to str via :func:`_cell_value_to_str` before any pandas dtype
    logic touches it.
    """
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=False)

    # Extract header row
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()

    headers: list[str] = []
    for i, cell in enumerate(header_row):
        raw_val: Any = None if isinstance(cell, MergedCell) else cell.value
        if raw_val is None or str(raw_val).strip() == "":
            headers.append(f"{_UNNAMED_PLACEHOLDER} {i}")
        else:
            headers.append(str(raw_val))

    # Build data rows
    data: list[list[str]] = []
    for row in rows_iter:
        row_data: list[str] = [
            "" if isinstance(cell, MergedCell) else _cell_value_to_str(cell.value)
            for cell in row
        ]
        data.append(row_data)

    # Pad or trim rows to match header width
    n_cols = len(headers)
    padded: list[list[str]] = [(r + [""] * n_cols)[:n_cols] for r in data]

    return pd.DataFrame(padded, columns=headers)


class ExcelParser:
    """Concrete XLSX/XLSM source-file parser (adapter for the SourceParser port).

    Uses openpyxl in read-only + data-only mode so that VBA macros are never
    loaded or executed, satisfying EB-06 and T-3.3 DoD.  All cell values are
    converted to strings before building the DataFrame, ensuring dtype=str
    universality and preserving Excel error sentinels such as '#N/A' (ADR-004).
    """

    def list_sheets(self, path: Path) -> list[str] | None:
        """Return the ordered list of sheet names without reading cell data.

        Args:
            path: Absolute path to the XLSX or XLSM file.

        Raises:
            UnsupportedFormatError: File extension not handled by this adapter.

        """
        if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
            msg = (
                f"ExcelParser does not handle '{path.suffix}' files. "
                f"Supported: {sorted(_SUPPORTED_EXTENSIONS)}"
            )
            raise UnsupportedFormatError(msg)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def parse(self, path: Path, **kwargs: object) -> ParsedSource:
        """Parse an Excel file and return a :class:`ParsedSource`.

        Args:
            path:        Absolute path to the XLSX or XLSM file.
            **kwargs:    Optional ``sheet_name`` (str) selects a specific sheet;
                         if omitted the first sheet is used.

        Raises:
            UnsupportedFormatError: File extension not handled by this adapter.

        """
        if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
            msg = (
                f"ExcelParser does not handle '{path.suffix}' files. "
                f"Supported: {sorted(_SUPPORTED_EXTENSIONS)}"
            )
            raise UnsupportedFormatError(msg)

        sheet_name_arg = kwargs.get("sheet_name")
        all_sheets = self.list_sheets(path) or []
        if sheet_name_arg is not None:
            target_sheet = str(sheet_name_arg)
        else:
            target_sheet = all_sheets[0] if all_sheets else ""

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        try:
            df = _read_sheet_as_dataframe(wb, target_sheet)
        finally:
            wb.close()

        sku_hint = _guess_sku_column([str(c) for c in df.columns])

        return ParsedSource(
            dataframe=df,
            sheet_name=target_sheet,
            detected_encoding=None,
            sku_column_hint=sku_hint,
        )
