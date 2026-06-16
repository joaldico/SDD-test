"""Ingestion API router — T-3.6/T-3.7/T-3.8/T-3.10 (gate CA-01) + T-4.2 (gate CA-03).

T-4.6: async process endpoint (ADR-002 TaskRunner) + status polling endpoint.

Deduplication (spec 2.6) and ``duplicate_findings`` persistence added in T-4.2.
Async pipeline via TaskRunner (ADR-002) and status polling added in T-4.6.

Endpoints:
  POST /api/v1/runs                              Create a new reconciliation run.
  POST /api/v1/runs/{run_id}/files               Upload a source file (+ save to staging).
  GET  /api/v1/runs/{run_id}/files/{file_id}/preview   Parse & preview a staged file.
  PUT  /api/v1/runs/{run_id}/files/{file_id}/mapping   Confirm column mapping.
  POST /api/v1/runs/{run_id}/process             Trigger reconciliation (gate RNF-08) → 202.
  GET  /api/v1/runs/{run_id}/status              Poll pipeline phase + summary_metrics.

Auth: uses the dummy ``get_current_user`` dependency (M2 bypass — see platform/deps.py).
DB:   uses the synchronous ``get_db`` session dependency (overridable in tests).

Staging: uploaded bytes are persisted to ``get_staging_dir()`` so the preview and
mapping endpoints can re-parse the same file without a second upload (RNF-05, T-3.7).
"""

from __future__ import annotations

import hashlib
import logging
import re
from collections.abc import Callable  # noqa: TC003
from datetime import UTC, datetime
from pathlib import Path
from typing import Annotated, Any

import openpyxl
import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import text as sa_text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session  # noqa: TC002

from marketplace_conciliator.ingestion.block_locator import BlockLocator, BlockNotFoundError
from marketplace_conciliator.ingestion.column_suggester import ColumnSuggester
from marketplace_conciliator.ingestion.csv_parser import CsvParser
from marketplace_conciliator.ingestion.deduplicator import (
    DuplicateFindingData,
    deduplicate_amazon_errors,
    deduplicate_feed,
    deduplicate_occ,
)
from marketplace_conciliator.ingestion.excel_parser import ExcelParser
from marketplace_conciliator.ingestion.sku_normalizer import normalise_sku
from marketplace_conciliator.platform.db.models.runs import (
    ColumnMapping,
    ReconciliationRun,
    SourceFile,
)
from marketplace_conciliator.platform.db.session import get_db
from marketplace_conciliator.platform.deps import CurrentUser, get_current_user
from marketplace_conciliator.reconciliation.error_classifier import ErrorRow
from marketplace_conciliator.reconciliation.persistence import RunItemData, persist_run_results
from marketplace_conciliator.reconciliation.ports import TaskRunnerPort  # noqa: TC001
from marketplace_conciliator.settings import get_settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/runs", tags=["ingestion"])


# ---------------------------------------------------------------------------
# Overridable dependencies — TaskRunner + DB factory (T-4.6, ADR-002)
# ---------------------------------------------------------------------------


def get_task_runner() -> TaskRunnerPort:
    """FastAPI dependency — returns the process-level TaskRunner.

    This stub is overridden by ``main.py`` (production wiring) and by tests
    (``_NoOpTaskRunner`` or ``_SyncTaskRunner``) via ``app.dependency_overrides``.
    """
    msg = "TaskRunner not wired — override get_task_runner in create_app() or test setup."
    raise RuntimeError(msg)  # pragma: no cover


def get_db_factory() -> Callable[[], Session]:
    """FastAPI dependency — returns a callable that creates a new DB session.

    Background work functions (running in the ThreadPool) must create their own
    sessions independently of the request lifecycle.  This dependency is
    overridden in tests to provide the in-memory SQLite session factory.
    """
    from marketplace_conciliator.platform.db.session import get_session_factory  # noqa: PLC0415

    return get_session_factory()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MAX_FILE_SIZE_BYTES: int = 50 * 1024 * 1024  # 50 MB

_VALID_ROLES: frozenset[str] = frozenset({"occ_top", "wm_feed", "amazon_report"})

_EXCEL_EXTENSIONS: frozenset[str] = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})
_CSV_EXTENSIONS: frozenset[str] = frozenset({".csv", ".txt", ".tsv"})

# Fraction of values that must be numeric for stock validation (mirrors column_suggester)
_STOCK_NUMERIC_THRESHOLD: float = 0.70
_NUMERIC_RE: re.Pattern[str] = re.compile(r"^[+-]?\d+([.,]\d+)?$")

# Number of sample rows returned in the preview response
_PREVIEW_SAMPLE_ROWS: int = 5


# ---------------------------------------------------------------------------
# Staging directory dependency
# ---------------------------------------------------------------------------


def get_staging_dir() -> Path:
    """Return the staging directory path, creating it if necessary.

    Overridable in tests via ``app.dependency_overrides[get_staging_dir]``.
    """
    d = get_settings().staging_dir
    d.mkdir(parents=True, exist_ok=True)
    return d


# ---------------------------------------------------------------------------
# Request / Response schemas — T-3.6
# ---------------------------------------------------------------------------


class CreateRunRequest(BaseModel):
    """Request body for creating a new reconciliation run."""

    marketplace: str = "amazon_es"


class RunResponse(BaseModel):
    """Response schema for a reconciliation run."""

    id: int
    user_id: int
    marketplace: str
    status: str
    created_at: datetime

    model_config = {"from_attributes": True}


class SourceFileResponse(BaseModel):
    """Response schema for an uploaded source file."""

    id: int
    run_id: int
    role: str
    original_filename: str
    sha256: str
    total_rows: int
    discarded_rows: int
    uploaded_at: datetime

    model_config = {"from_attributes": True}


# ---------------------------------------------------------------------------
# Request / Response schemas — T-3.7 (Preview contract, plan 3.7)
# ---------------------------------------------------------------------------


class SheetInfo(BaseModel):
    """Sheet name and approximate row count for an Excel workbook."""

    name: str
    rows: int


class BlockInfo(BaseModel):
    """Location metadata for a structured block inside a sheet (EB-02/03)."""

    title_matched: str
    header_row: int
    data_start_row: int


class HeaderInfo(BaseModel):
    """One column header descriptor."""

    index: int
    name: str
    technical_name: str | None = None


class ColumnSuggestionSchema(BaseModel):
    """Serializable form of a :class:`ColumnSuggestion`."""

    column_index: int
    confidence: float
    reason: str


class PreviewWarning(BaseModel):
    """Non-fatal issue encountered during preview parsing."""

    code: str
    message: str
    row: int | None = None


class PreviewResponse(BaseModel):
    """Contract for GET .../preview — plan 3.7."""

    file_role: str
    sheet: str | None
    available_sheets: list[SheetInfo] | None
    block: BlockInfo | None
    headers: list[HeaderInfo]
    sample_rows: list[list[str]]
    suggestions: dict[str, ColumnSuggestionSchema]
    warnings: list[PreviewWarning]
    discarded_rows: int


# ---------------------------------------------------------------------------
# Request / Response schemas — T-3.8 (Mapping)
# ---------------------------------------------------------------------------


class MappingItem(BaseModel):
    """A single logical-field → physical-column mapping."""

    logical_field: str
    column_index: int
    was_suggested: bool = False


class MappingRequest(BaseModel):
    """Request body for PUT .../mapping."""

    mappings: list[MappingItem]


class MappingWarning(BaseModel):
    """Warning issued during mapping validation."""

    code: str
    message: str
    sample: list[str] | None = None


class MappingResponse(BaseModel):
    """Response for PUT .../mapping."""

    status: str  # "ok" | "warnings"
    warnings: list[MappingWarning]


# ---------------------------------------------------------------------------
# Endpoints — T-3.6
# ---------------------------------------------------------------------------


@router.post(
    "",
    status_code=status.HTTP_201_CREATED,
    response_model=RunResponse,
    summary="Create a new reconciliation run",
)
def create_run(
    body: CreateRunRequest,
    current_user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
) -> RunResponse:
    """Create a new reconciliation run and return its metadata."""
    run = ReconciliationRun(
        user_id=current_user.id,
        marketplace=body.marketplace,
        status="uploaded",
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    return RunResponse(
        id=run.id,
        user_id=run.user_id,
        marketplace=run.marketplace,
        status=run.status,
        created_at=run.created_at,
    )


@router.post(
    "/{run_id}/files",
    status_code=status.HTTP_201_CREATED,
    response_model=SourceFileResponse,
    summary="Upload a source file for a run",
)
def upload_file(  # noqa: PLR0913
    run_id: int,
    role: Annotated[str, Form(description="one of: occ_top, wm_feed, amazon_report")],
    file: Annotated[UploadFile, File(description="The source file to upload")],
    current_user: Annotated[CurrentUser, Depends(get_current_user)],  # noqa: ARG001
    db: Annotated[Session, Depends(get_db)],
    staging_dir: Annotated[Path, Depends(get_staging_dir)],
) -> SourceFileResponse:
    """Upload a source file and attach it to an existing run.

    Validates:
    - Run exists (404 if not found).
    - Role is one of the valid values (422).
    - File size ≤ 50 MB (413).
    - No duplicate role for the same run (409).

    Side-effect: raw bytes are saved to staging_dir/{id}{ext} so that the
    preview and mapping endpoints can re-parse without a second upload.
    """
    if role not in _VALID_ROLES:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid role '{role}'. Must be one of: {sorted(_VALID_ROLES)}",
        )

    run = db.get(ReconciliationRun, run_id)
    if run is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Run {run_id} not found.",
        )

    raw_bytes: bytes = file.file.read()
    if len(raw_bytes) > _MAX_FILE_SIZE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=(
                f"File size {len(raw_bytes):,} bytes exceeds the 50 MB limit "
                f"({_MAX_FILE_SIZE_BYTES:,} bytes)."
            ),
        )

    sha256_hex: str = hashlib.sha256(raw_bytes).hexdigest()
    original_filename = file.filename or "unknown"

    source_file = SourceFile(
        run_id=run_id,
        role=role,
        original_filename=original_filename,
        sha256=sha256_hex,
        detected_encoding=None,
        detected_delimiter=None,
        sheet_name=None,
        data_start_row=None,
        total_rows=0,
        discarded_rows=0,
        uploaded_at=datetime.now(tz=UTC),
    )
    db.add(source_file)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"A file with role '{role}' already exists for run {run_id}.",
        ) from None

    db.refresh(source_file)

    # Persist raw bytes to staging volume (enables preview / mapping without re-upload)
    ext = Path(original_filename).suffix.lower()
    staging_path = staging_dir / f"{source_file.id}{ext}"
    staging_path.write_bytes(raw_bytes)

    return SourceFileResponse(
        id=source_file.id,
        run_id=source_file.run_id,
        role=source_file.role,
        original_filename=source_file.original_filename,
        sha256=source_file.sha256,
        total_rows=source_file.total_rows,
        discarded_rows=source_file.discarded_rows,
        uploaded_at=source_file.uploaded_at,
    )


# ---------------------------------------------------------------------------
# Endpoint — T-3.7: GET preview
# ---------------------------------------------------------------------------


@router.get(
    "/{run_id}/files/{file_id}/preview",
    response_model=PreviewResponse,
    summary="Preview a staged source file with column suggestions",
)
def preview_file(
    run_id: int,
    file_id: int,
    db: Annotated[Session, Depends(get_db)],
    staging_dir: Annotated[Path, Depends(get_staging_dir)],
    sheet: Annotated[str | None, Query(description="Sheet name to preview (Excel only)")] = None,
) -> PreviewResponse:
    """Parse and preview a previously uploaded source file.

    Returns the exact preview contract from plan 3.7:
    - available_sheets with row counts (Excel only)
    - block location info (amazon_report only)
    - column headers and sample data rows
    - heuristic column suggestions with confidence + reason (OBJ-03)
    - warnings for non-fatal issues (discarded rows, missing blocks, etc.)
    """
    run = db.get(ReconciliationRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")

    source_file = db.get(SourceFile, file_id)
    if source_file is None or source_file.run_id != run_id:
        raise HTTPException(status_code=404, detail=f"File {file_id} not found for run {run_id}.")

    ext = Path(source_file.original_filename).suffix.lower()
    staging_path = staging_dir / f"{file_id}{ext}"
    if not staging_path.exists():
        raise HTTPException(
            status_code=404,
            detail=(
                f"Staged file for source_file {file_id} not found. "
                "Re-upload the file to regenerate the preview."
            ),
        )

    if ext in _EXCEL_EXTENSIONS:
        return _preview_excel(source_file, staging_path, sheet)
    if ext in _CSV_EXTENSIONS:
        return _preview_csv(source_file, staging_path)

    raise HTTPException(
        status_code=422,
        detail=f"Unsupported file extension '{ext}' for preview.",
    )


# ---------------------------------------------------------------------------
# Endpoint — T-3.8: PUT mapping
# ---------------------------------------------------------------------------


@router.put(
    "/{run_id}/files/{file_id}/mapping",
    response_model=MappingResponse,
    summary="Confirm column mapping for a source file",
)
def create_mapping(  # noqa: PLR0913
    run_id: int,
    file_id: int,
    body: MappingRequest,
    current_user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    staging_dir: Annotated[Path, Depends(get_staging_dir)],
) -> MappingResponse:
    """Confirm the column mapping for a source file and persist it.

    Validates:
    - Run exists (404).
    - File belongs to the run (404).
    - Stock column has ≥ 70% numeric values — if not, issues STOCK_NOT_NUMERIC warning
      but still persists the mapping (CA-04: degradación explícita, no hard rejection).
    - Out-of-range column indices: issues INVALID_COLUMN_INDEX warning, skips that entry.

    Upsert behaviour: a second PUT for the same (source_file_id, logical_field) replaces
    the previous row (UNIQUE constraint satisfied by delete-before-insert).

    Side-effect: run status advances to 'mapping' if still 'uploaded'.
    """
    run = db.get(ReconciliationRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")

    source_file = db.get(SourceFile, file_id)
    if source_file is None or source_file.run_id != run_id:
        raise HTTPException(status_code=404, detail=f"File {file_id} not found for run {run_id}.")

    # Load the DataFrame for validation (same parse logic as preview)
    ext = Path(source_file.original_filename).suffix.lower()
    staging_path = staging_dir / f"{file_id}{ext}"
    if not staging_path.exists():
        raise HTTPException(
            status_code=404,
            detail=(
                f"Staged file for source_file {file_id} not found. "
                "Re-upload the file before mapping."
            ),
        )

    df = _load_dataframe(source_file, staging_path)
    n_cols = len(df.columns)

    warnings: list[MappingWarning] = []
    now = datetime.now(tz=UTC)

    for item in body.mappings:
        # Validate column index range
        if item.column_index < 0 or item.column_index >= n_cols:
            warnings.append(
                MappingWarning(
                    code="INVALID_COLUMN_INDEX",
                    message=(
                        f"Column index {item.column_index} is out of range "
                        f"(file has {n_cols} columns)."
                    ),
                ),
            )
            continue

        col_name = str(df.columns[item.column_index])

        # Validate numeric ratio for stock fields (CA-04)
        if item.logical_field == "stock":
            ratio = _numeric_ratio(df, col_name)
            if ratio < _STOCK_NUMERIC_THRESHOLD:
                sample = (
                    df[col_name].dropna().astype(str).head(3).tolist()
                    if col_name in df.columns
                    else []
                )
                warnings.append(
                    MappingWarning(
                        code="STOCK_NOT_NUMERIC",
                        message=(
                            f"Column '{col_name}' has only {ratio:.0%} numeric values. "
                            "Stock quantities may be incorrect (degraded mode)."
                        ),
                        sample=sample,
                    ),
                )

        # Upsert: remove existing mapping for (source_file_id, logical_field) before inserting
        existing = (
            db.query(ColumnMapping)
            .filter(
                ColumnMapping.source_file_id == file_id,
                ColumnMapping.logical_field == item.logical_field,
            )
            .first()
        )
        if existing is not None:
            db.delete(existing)
            db.flush()

        db.add(
            ColumnMapping(
                source_file_id=file_id,
                logical_field=item.logical_field,
                source_column_name=col_name,
                source_column_index=item.column_index,
                was_suggested=item.was_suggested,
                confirmed_by=current_user.id,
                confirmed_at=now,
            ),
        )

    # Advance run status: uploaded → mapping (first time any mapping is confirmed)
    if run.status == "uploaded" and body.mappings:
        run.status = "mapping"

    db.commit()

    return MappingResponse(
        status="warnings" if warnings else "ok",
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Endpoint — T-3.10 gate CA-01 / T-4.2 gate CA-03
# ---------------------------------------------------------------------------


class ProcessResponse(BaseModel):
    """202 response body for POST /runs/{run_id}/process."""

    status_url: str


class RunStatusResponse(BaseModel):
    """Response body for GET /runs/{run_id}/status (ADR-002 polling contract)."""

    status: str
    phase: str | None
    failure_reason: str | None
    summary_metrics: dict[str, Any] | None


# ---------------------------------------------------------------------------
# Endpoint — T-4.6: GET /runs/{run_id}/status (polling, ADR-002)
# ---------------------------------------------------------------------------


@router.get(
    "/{run_id}/status",
    response_model=RunStatusResponse,
    summary="Poll the pipeline phase and summary metrics for a run (T-4.6)",
)
def get_run_status(
    run_id: int,
    db: Annotated[Session, Depends(get_db)],
) -> RunStatusResponse:
    """Return current status, pipeline phase, failure_reason, and summary_metrics.

    Used by the frontend polling loop (every 2 s) to track pipeline progress
    from 'processing' through to 'completed' or 'failed' (plan 3.5, ADR-002).
    """
    run = db.get(ReconciliationRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")

    metrics: dict[str, Any] | None = None
    if run.summary_metrics is not None:
        # summary_metrics is stored as JSON; SQLAlchemy deserialises to dict for MySQL,
        # but SQLite tests store it as a raw JSON string — handle both.
        if isinstance(run.summary_metrics, str):
            import json  # noqa: PLC0415

            metrics = json.loads(run.summary_metrics)
        else:
            metrics = run.summary_metrics

    return RunStatusResponse(
        status=run.status,
        phase=run.phase,
        failure_reason=run.failure_reason,
        summary_metrics=metrics,
    )


# Logical field names for Amazon error columns (used by deduplicator)
_AMAZON_ERROR_KEY_FIELDS: tuple[str, ...] = (
    "error_code",
    "error_message",
    "affected_field",
)

# Recognised column name fragments for Amazon error detection (fallback heuristic)
_ERROR_CODE_HINTS: tuple[str, ...] = ("código", "code", "error_code")
_ERROR_CAT_HINTS: tuple[str, ...] = ("categoría", "category", "error_category")
_ERROR_MSG_HINTS: tuple[str, ...] = ("mensaje", "message", "error_message")
_AFFECTED_FIELD_HINTS: tuple[str, ...] = ("campo", "affected", "affected_field")


@router.post(
    "/{run_id}/process",
    status_code=202,
    response_model=ProcessResponse,
    summary="Trigger async reconciliation processing (gate RNF-08, ADR-002, T-4.6)",
)
def trigger_process(  # noqa: PLR0913
    run_id: int,
    current_user: Annotated[CurrentUser, Depends(get_current_user)],  # noqa: ARG001
    db: Annotated[Session, Depends(get_db)],
    staging_dir: Annotated[Path, Depends(get_staging_dir)],
    task_runner: Annotated[TaskRunnerPort, Depends(get_task_runner)],
    db_factory: Annotated[Callable[[], Session], Depends(get_db_factory)],
) -> ProcessResponse:
    """Gate check (RNF-08) → 409 if mapping incomplete, 202 + enqueue otherwise.

    On success (202, ADR-002):
      1.  Verifies all 3 required files have a confirmed SKU mapping (gate).
      2.  Transitions run status to 'processing' and commits immediately.
      3.  Submits the full reconciliation pipeline to the TaskRunner.
          The caller receives 202 before the pipeline starts.
      4.  The frontend polls GET /runs/{id}/status every 2 s until completed.

    Pipeline stages (run inside the TaskRunner thread):
      Validando → Normalizando → Deduplicando → Cruzando → Persistiendo → Listo
    """
    run = db.get(ReconciliationRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")

    files = db.query(SourceFile).filter(SourceFile.run_id == run_id).all()
    file_by_role: dict[str, SourceFile] = {sf.role: sf for sf in files}

    required_roles: frozenset[str] = frozenset({"occ_top", "wm_feed", "amazon_report"})
    missing_roles = required_roles - set(file_by_role.keys())
    if missing_roles:
        raise HTTPException(
            status_code=409,
            detail=(
                f"mapeo pendiente de confirmación: "
                f"ficheros no cargados: {', '.join(sorted(missing_roles))}"
            ),
        )

    # ── Gate: SKU mapping must be confirmed for every file (RNF-08) ─────────
    for role, source_file in file_by_role.items():
        sku_mapping = (
            db.query(ColumnMapping)
            .filter(
                ColumnMapping.source_file_id == source_file.id,
                ColumnMapping.logical_field == "sku",
            )
            .first()
        )
        if sku_mapping is None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"mapeo pendiente de confirmación: "
                    f"fichero '{role}' sin campo SKU confirmado"
                ),
            )

    # Gate passed: transition to processing and commit before returning 202.
    run.status = "processing"
    db.commit()

    # Capture staging_dir in the closure (Path is immutable — safe to share).
    _staging = staging_dir

    def _work_fn(rid: int) -> None:
        """Full reconciliation pipeline — executes in a ThreadPool worker thread."""
        _execute_pipeline(rid, db_factory, _staging)

    task_runner.submit(run_id, _work_fn)

    return ProcessResponse(status_url=f"/api/v1/runs/{run_id}/status")


# ---------------------------------------------------------------------------
# Pipeline execution — T-4.6 (extracted from trigger_process for async use)
# ---------------------------------------------------------------------------


def _update_phase(db: Session, run_id: int, phase: str) -> None:
    """Write the pipeline phase to reconciliation_runs for polling visibility."""
    try:
        db.execute(
            sa_text("UPDATE reconciliation_runs SET phase = :phase WHERE id = :id"),
            {"phase": phase, "id": run_id},
        )
        db.commit()
    except Exception:  # noqa: BLE001
        logger.warning("_update_phase: failed to update phase for run %d", run_id)


def _execute_pipeline(  # noqa: C901, PLR0912, PLR0915
    run_id: int,
    db_factory: Callable[[], Session],
    staging_dir: Path,
) -> None:
    """Full reconciliation pipeline for run_id — runs inside a ThreadPool thread.

    Creates its own DB session from ``db_factory`` to be independent of the
    request lifecycle (ADR-002).  Phase transitions are persisted directly so
    that the polling endpoint reflects progress in real time (plan 3.4).

    Stages (plan 3.4):
      Validando → Deduplicando → Cruzando → Persistiendo → Listo
    """
    db = db_factory()
    try:
        _update_phase(db, run_id, "Validando")

        files = db.query(SourceFile).filter(SourceFile.run_id == run_id).all()
        file_by_role: dict[str, SourceFile] = {sf.role: sf for sf in files}

        _update_phase(db, run_id, "Deduplicando")

        run_item_data: dict[str, _RunItemAccum] = {}
        amazon_error_rows: list[dict[str, object]] = []

        for role, source_file in file_by_role.items():
            all_mappings: dict[str, ColumnMapping] = {
                cm.logical_field: cm
                for cm in db.query(ColumnMapping)
                .filter(ColumnMapping.source_file_id == source_file.id)
                .all()
            }

            sku_mapping = all_mappings.get("sku")
            if sku_mapping is None:
                continue

            ext = Path(source_file.original_filename).suffix.lower()
            staging_path = staging_dir / f"{source_file.id}{ext}"
            if not staging_path.exists():
                continue

            df = _load_dataframe(source_file, staging_path)
            if sku_mapping.source_column_index >= len(df.columns):
                continue

            sku_col = str(df.columns[sku_mapping.source_column_index])

            if role == "occ_top":
                dedup_result = deduplicate_occ(df, sku_col)
                stock_mapping = all_mappings.get("stock")
                stock_col_name = (
                    str(df.columns[stock_mapping.source_column_index])
                    if stock_mapping and stock_mapping.source_column_index < len(df.columns)
                    else None
                )
                _persist_findings(db, source_file.id, dedup_result.findings)
                for _, row in dedup_result.dataframe.iterrows():
                    raw_str = str(row[sku_col]) if row[sku_col] is not None else ""
                    norm = normalise_sku(raw_str)
                    if not norm.is_valid or norm.value is None:
                        continue
                    sku_n = norm.value
                    occ_stock: int | None = None
                    if stock_col_name and stock_col_name in row.index:
                        occ_stock = _safe_int(row[stock_col_name])
                    accum = run_item_data.setdefault(
                        sku_n, _RunItemAccum(sku_norm=sku_n, sku_raw=raw_str),
                    )
                    accum.in_occ = True
                    if occ_stock is not None:
                        accum.occ_stock = occ_stock
                    if not accum.sku_raw:
                        accum.sku_raw = raw_str

            elif role == "wm_feed":
                stock_mapping = all_mappings.get("stock")
                stock_col_name = (
                    str(df.columns[stock_mapping.source_column_index])
                    if stock_mapping and stock_mapping.source_column_index < len(df.columns)
                    else None
                )
                dedup_result = deduplicate_feed(df, sku_col, stock_col_name)
                _persist_findings(db, source_file.id, dedup_result.findings)
                for _, row in dedup_result.dataframe.iterrows():
                    raw_str = str(row[sku_col]) if row[sku_col] is not None else ""
                    norm = normalise_sku(raw_str)
                    if not norm.is_valid or norm.value is None:
                        continue
                    sku_n = norm.value
                    feed_stock: int | None = None
                    if stock_col_name and stock_col_name in row.index:
                        feed_stock = _safe_int(row[stock_col_name])
                    accum = run_item_data.setdefault(
                        sku_n, _RunItemAccum(sku_norm=sku_n, sku_raw=raw_str),
                    )
                    accum.in_feed = True
                    if feed_stock is not None:
                        accum.feed_stock = feed_stock
                    if sku_n in dedup_result.stock_conflicts:
                        accum.stock_conflict = True
                    if not accum.sku_raw:
                        accum.sku_raw = raw_str

            elif role == "amazon_report":
                error_col_names = _resolve_amazon_error_cols(df, all_mappings)
                dedup_result = deduplicate_amazon_errors(
                    df, sku_col, error_col_names.error_key_cols,
                )
                _persist_findings(db, source_file.id, dedup_result.findings)
                for _, row in dedup_result.dataframe.iterrows():
                    raw_str = str(row[sku_col]) if row[sku_col] is not None else ""
                    norm = normalise_sku(raw_str)
                    if not norm.is_valid or norm.value is None:
                        continue
                    sku_n = norm.value
                    accum = run_item_data.setdefault(
                        sku_n, _RunItemAccum(sku_norm=sku_n, sku_raw=raw_str),
                    )
                    accum.in_amazon = True
                    if not accum.sku_raw:
                        accum.sku_raw = raw_str
                    raw_error_code = str(
                        row.get(error_col_names.error_code_col, "") or "",
                    ).strip()
                    if not raw_error_code:
                        continue
                    amazon_error_rows.append({
                        "sku_norm": sku_n,
                        "error_code": raw_error_code,
                        "error_category": str(
                            row.get(error_col_names.error_category_col, "ERROR") or "ERROR",
                        ),
                        "error_message": str(
                            row.get(error_col_names.error_message_col, "") or "",
                        ),
                        "affected_field": str(
                            row.get(error_col_names.affected_field_col, "") or "",
                        ) or None,
                    })

        db.flush()

        _update_phase(db, run_id, "Cruzando")

        skus_with_errors: set[str] = {str(row["sku_norm"]) for row in amazon_error_rows}

        errors_by_sku: dict[str, list[ErrorRow]] = {}
        for err_dict in amazon_error_rows:
            sku_n = str(err_dict["sku_norm"])
            errors_by_sku.setdefault(sku_n, []).append(
                ErrorRow(
                    sku_norm=sku_n,
                    error_code=str(err_dict["error_code"] or "UNKNOWN"),
                    error_category=str(err_dict["error_category"] or "ERROR"),
                    error_message=str(err_dict["error_message"] or ""),
                    affected_field=(
                        str(err_dict["affected_field"])
                        if err_dict.get("affected_field")
                        else None
                    ),
                ),
            )

        run_items_data: list[RunItemData] = []
        for sku_n, accum in run_item_data.items():
            sync_status = _compute_sync_status(
                _in_occ=accum.in_occ,
                in_feed=accum.in_feed,
                in_amazon=accum.in_amazon,
                has_errors=sku_n in skus_with_errors,
            )
            run_items_data.append(
                RunItemData(
                    sku_norm=sku_n,
                    sku_raw=accum.sku_raw,
                    in_occ=accum.in_occ,
                    in_feed=accum.in_feed,
                    in_amazon_report=accum.in_amazon,
                    sync_status=sync_status,
                    feed_stock=accum.feed_stock,
                    occ_stock=accum.occ_stock,
                    stock_conflict=accum.stock_conflict,
                    error_rows=errors_by_sku.get(sku_n, []),
                ),
            )

        _update_phase(db, run_id, "Persistiendo")
        persist_run_results(db, run_id, run_items_data)

    except Exception:
        logger.exception("Pipeline failed for run %d", run_id)
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Endpoint — T-4.3: GET catalog-health (Vista 3 minimal, spec 2.7)
# ---------------------------------------------------------------------------


class CatalogHealthItem(BaseModel):
    """One run_item row for the catalog-health view."""

    sku_norm: str
    sku_raw: str
    sync_status: str
    feed_stock: int | None
    occ_stock: int | None
    stock_conflict: bool
    in_occ: bool
    in_feed: bool
    in_amazon_report: bool
    stock_disponible: bool  # feed_stock > 0


class CatalogHealthResponse(BaseModel):
    """Response for GET /runs/{run_id}/catalog-health."""

    run_id: int
    items: list[CatalogHealthItem]
    total: int


@router.get(
    "/{run_id}/catalog-health",
    response_model=CatalogHealthResponse,
    summary="Vista 3 — Salud de catálogo (spec 2.7, T-4.3)",
)
def catalog_health(
    run_id: int,
    db: Annotated[Session, Depends(get_db)],
    sync_status: Annotated[
        str | None,
        Query(description="Filter by sync_status (e.g. DESYNC_FEED_ONLY)"),
    ] = None,
) -> CatalogHealthResponse:
    """Return run_items ordered by priority (sync_status) then feed_stock DESC.

    Default ordering follows spec 2.7: highest-priority status first,
    within each status ordered by feed_stock DESC (active stock loss first).
    Optional ``sync_status`` query param filters to a single classification.
    """
    run = db.get(ReconciliationRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")

    query = sa_text("""
        SELECT sku_norm, sku_raw, sync_status,
               feed_stock, occ_stock, stock_conflict,
               in_occ, in_feed, in_amazon_report
        FROM run_items
        WHERE run_id = :run_id
          AND (:sync_status IS NULL OR sync_status = :sync_status)
        ORDER BY
            CASE sync_status
                WHEN 'SENT_WITH_ERROR'    THEN 1
                WHEN 'DESYNC_FEED_ONLY'   THEN 2
                WHEN 'DESYNC_AMAZON_ONLY' THEN 3
                WHEN 'NOT_SENT'           THEN 4
                WHEN 'SENT_OK'            THEN 5
                ELSE 9
            END ASC,
            COALESCE(feed_stock, 0) DESC
    """)

    rows = db.execute(query, {"run_id": run_id, "sync_status": sync_status}).fetchall()

    items = [
        CatalogHealthItem(
            sku_norm=str(row[0]),
            sku_raw=str(row[1]),
            sync_status=str(row[2]),
            feed_stock=row[3],
            occ_stock=row[4],
            stock_conflict=bool(row[5]),
            in_occ=bool(row[6]),
            in_feed=bool(row[7]),
            in_amazon_report=bool(row[8]),
            stock_disponible=(row[3] is not None and row[3] > 0),
        )
        for row in rows
    ]

    return CatalogHealthResponse(run_id=run_id, items=items, total=len(items))


# ---------------------------------------------------------------------------
# Processing helpers — T-4.3: 3-way cross-join + sync_status computation
# ---------------------------------------------------------------------------


def _compute_sync_status(
    _in_occ: bool,  # noqa: FBT001  — kept for call-site clarity (spec 2.7 table)
    in_feed: bool,  # noqa: FBT001
    in_amazon: bool,  # noqa: FBT001
    has_errors: bool,  # noqa: FBT001
) -> str:
    """Assign one of the 5 sync_status values per spec 2.7.

    Priority matrix (∈ occ / ∈ feed / ∈ amazon):
      feed ∧ amazon ∧ errors  → SENT_WITH_ERROR
      feed ∧ amazon ∧ ¬errors → SENT_OK
      feed ∧ ¬amazon          → DESYNC_FEED_ONLY
      ¬feed ∧ amazon          → DESYNC_AMAZON_ONLY
      occ ∧ ¬feed ∧ ¬amazon  → NOT_SENT  (fallback)
    """
    if in_feed and in_amazon:
        return "SENT_WITH_ERROR" if has_errors else "SENT_OK"
    if in_feed:
        return "DESYNC_FEED_ONLY"
    if in_amazon:
        return "DESYNC_AMAZON_ONLY"
    return "NOT_SENT"


# ---------------------------------------------------------------------------
# Processing helpers — T-4.2
# ---------------------------------------------------------------------------


class _RunItemAccum:
    """Mutable accumulator for a single sku_norm across all 3 source files."""

    __slots__ = (
        "feed_stock",
        "in_amazon",
        "in_feed",
        "in_occ",
        "occ_stock",
        "sku_norm",
        "sku_raw",
        "stock_conflict",
    )

    def __init__(self, sku_norm: str, sku_raw: str) -> None:
        """Initialise all flags to False / None."""
        self.sku_norm = sku_norm
        self.sku_raw = sku_raw
        self.in_occ: bool = False
        self.in_feed: bool = False
        self.in_amazon: bool = False
        self.feed_stock: int | None = None
        self.occ_stock: int | None = None
        self.stock_conflict: bool = False


def _persist_findings(
    db: Session,
    source_file_id: int,
    findings: list[DuplicateFindingData],
) -> None:
    """Insert duplicate_findings rows for every finding in the list.

    Uses raw SQL so it works on both SQLite (tests) and MySQL (production).
    The INSERT OR IGNORE prevents double-inserts if the pipeline is called
    twice for the same run (defensive idempotency).
    """
    for fd in findings:
        import json  # noqa: PLC0415
        db.execute(
            sa_text("""
                INSERT OR IGNORE INTO duplicate_findings
                    (source_file_id, sku_norm, occurrences, resolution, discarded_values)
                VALUES
                    (:source_file_id, :sku_norm, :occurrences, :resolution,
                     :discarded_values)
            """),
            {
                "source_file_id": source_file_id,
                "sku_norm": fd.sku_norm,
                "occurrences": fd.occurrences,
                "resolution": fd.resolution,
                "discarded_values": json.dumps(fd.discarded_values),
            },
        )


def _safe_int(value: object) -> int | None:
    """Parse a cell value to int, returning None for non-numeric input."""
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


class _AmazonErrorCols:
    """Column name resolution result for Amazon report error fields."""

    __slots__ = (
        "affected_field_col",
        "error_category_col",
        "error_code_col",
        "error_key_cols",
        "error_message_col",
    )

    def __init__(
        self,
        error_code_col: str | None,
        error_category_col: str | None,
        error_message_col: str | None,
        affected_field_col: str | None,
        error_key_cols: list[str],
    ) -> None:
        self.error_code_col = error_code_col
        self.error_category_col = error_category_col
        self.error_message_col = error_message_col
        self.affected_field_col = affected_field_col
        self.error_key_cols = error_key_cols


def _resolve_amazon_error_cols(
    df: pd.DataFrame,
    all_mappings: dict[str, ColumnMapping],
) -> _AmazonErrorCols:
    """Return column names for Amazon error fields.

    Prefers confirmed ``column_mappings``.  Falls back to heuristic name
    matching against the DataFrame columns.
    """

    def _col_name_from_mapping(field: str) -> str | None:
        m = all_mappings.get(field)
        if m and m.source_column_index < len(df.columns):
            return str(df.columns[m.source_column_index])
        return None

    def _heuristic(hints: tuple[str, ...]) -> str | None:
        for col in df.columns:
            col_l = str(col).lower()
            if any(h in col_l for h in hints):
                return str(col)
        return None

    error_code_col = _col_name_from_mapping("error_code") or _heuristic(
        _ERROR_CODE_HINTS,
    )
    error_cat_col = _col_name_from_mapping("error_category") or _heuristic(
        _ERROR_CAT_HINTS,
    )
    error_msg_col = _col_name_from_mapping("error_message") or _heuristic(
        _ERROR_MSG_HINTS,
    )
    affected_col = _col_name_from_mapping("affected_field") or _heuristic(
        _AFFECTED_FIELD_HINTS,
    )

    error_key_cols: list[str] = [
        c for c in [error_code_col, error_msg_col, affected_col] if c
    ]

    return _AmazonErrorCols(
        error_code_col=error_code_col,
        error_category_col=error_cat_col,
        error_message_col=error_msg_col,
        affected_field_col=affected_col,
        error_key_cols=error_key_cols,
    )


# ---------------------------------------------------------------------------
# Private parsing helpers
# ---------------------------------------------------------------------------


def _staging_path(staging_dir: Path, file_id: int, original_filename: str) -> Path:
    ext = Path(original_filename).suffix.lower()
    return staging_dir / f"{file_id}{ext}"


def _count_sheet_rows(path: Path) -> dict[str, int]:
    """Return {sheet_name: max_row} for all sheets without reading cell data."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        return {name: wb[name].max_row or 0 for name in wb.sheetnames}
    finally:
        wb.close()


def _default_sheet_for_role(role: str, all_sheets: list[str]) -> str | None:
    """Pick the most relevant sheet for a given file role heuristically."""
    if not all_sheets:
        return None
    role_lower = role.lower()
    hints: list[str]
    if "amazon_report" in role_lower:
        hints = ["resumen", "summary", "processing"]
    elif "occ_top" in role_lower:
        hints = ["plantilla", "template", "datos", "data"]
    else:
        hints = []
    for hint in hints:
        for name in all_sheets:
            if hint in name.lower():
                return name
    return all_sheets[0]


def _build_suggestions(df: pd.DataFrame) -> dict[str, ColumnSuggestionSchema]:
    """Run ColumnSuggester and return the top suggestion per logical field."""
    suggestions_list = ColumnSuggester().suggest(df)
    best: dict[str, ColumnSuggestionSchema] = {}
    for s in suggestions_list:
        key = s.logical_field.value
        if key not in best or s.confidence > best[key].confidence:
            best[key] = ColumnSuggestionSchema(
                column_index=s.column_index,
                confidence=s.confidence,
                reason=s.reason,
            )
    return best


def _dataframe_to_sample(df: pd.DataFrame, n: int = _PREVIEW_SAMPLE_ROWS) -> list[list[str]]:
    """Convert the first n rows of a DataFrame to a list-of-lists of strings."""
    return [
        [str(v) if v is not None else "" for v in row]
        for row in df.head(n).to_numpy(dtype=object).tolist()
    ]


def _headers_from_df(df: pd.DataFrame) -> list[HeaderInfo]:
    return [HeaderInfo(index=i, name=str(col)) for i, col in enumerate(df.columns)]


def _numeric_ratio(df: pd.DataFrame, col_name: str) -> float:
    """Return fraction of non-empty values in col_name that match a numeric pattern."""
    try:
        series = df[col_name].dropna().astype(str)
        series = series[series.str.strip() != ""]
        if len(series) == 0:
            return 0.0
        count = series.apply(lambda v: bool(_NUMERIC_RE.match(v.strip()))).sum()
        return float(count) / len(series)
    except Exception:  # noqa: BLE001
        return 0.0


def _preview_csv(source_file: SourceFile, path: Path) -> PreviewResponse:
    """Build preview response for a flat CSV/TXT file."""
    parsed = CsvParser().parse(path)
    df = parsed.dataframe
    suggestions = _build_suggestions(df)

    return PreviewResponse(
        file_role=source_file.role,
        sheet=None,
        available_sheets=None,
        block=None,
        headers=_headers_from_df(df),
        sample_rows=_dataframe_to_sample(df),
        suggestions=suggestions,
        warnings=[],
        discarded_rows=0,
    )


def _preview_excel(
    source_file: SourceFile,
    path: Path,
    sheet_param: str | None,
) -> PreviewResponse:
    """Build preview response for an XLSX/XLSM file."""
    excel_parser = ExcelParser()
    all_sheet_names: list[str] = excel_parser.list_sheets(path) or []

    # Row counts (fast — reads worksheet dimension XML, not all cells)
    row_counts: dict[str, int] = {}
    try:
        row_counts = _count_sheet_rows(path)
    except Exception:  # noqa: BLE001
        row_counts = dict.fromkeys(all_sheet_names, 0)

    available_sheets = [
        SheetInfo(name=name, rows=row_counts.get(name, 0))
        for name in all_sheet_names
    ]

    # Determine target sheet
    target_sheet: str = (
        sheet_param
        if sheet_param and sheet_param in all_sheet_names
        else (_default_sheet_for_role(source_file.role, all_sheet_names) or "")
    )

    parsed = excel_parser.parse(path, sheet_name=target_sheet)
    df = parsed.dataframe

    block_info: BlockInfo | None = None
    preview_warnings: list[PreviewWarning] = []
    discarded = 0

    role = source_file.role
    locator = BlockLocator()

    if role == "amazon_report":
        try:
            located = locator.locate_errors_block(df)
            df = located.dataframe
            discarded = located.discarded_rows
            block_info = BlockInfo(
                title_matched="Errores y advertencias por SKU",
                header_row=located.title_row,
                data_start_row=located.data_start_row,
            )
            preview_warnings.extend(
                PreviewWarning(code="BLOCK_WARNING", message=w) for w in located.warnings
            )
        except BlockNotFoundError as exc:
            preview_warnings.append(
                PreviewWarning(code="BLOCK_NOT_FOUND", message=str(exc)),
            )

    elif role == "occ_top" and "plantilla" in target_sheet.lower():
        located = locator.parse_plantilla(df)
        df = located.dataframe
        discarded = located.discarded_rows
        if discarded > 0:
            preview_warnings.append(
                PreviewWarning(
                    code="EXAMPLE_ROW_DISCARDED",
                    message=(
                        f"Fila(s) de ejemplo de Amazon descartada(s) (EB-04): {discarded} fila(s)."
                    ),
                    row=6,
                ),
            )

    suggestions = _build_suggestions(df)

    return PreviewResponse(
        file_role=source_file.role,
        sheet=target_sheet,
        available_sheets=available_sheets,
        block=block_info,
        headers=_headers_from_df(df),
        sample_rows=_dataframe_to_sample(df),
        suggestions=suggestions,
        warnings=preview_warnings,
        discarded_rows=discarded,
    )


def _load_dataframe(source_file: SourceFile, path: Path) -> pd.DataFrame:
    """Parse the staged file and return the relevant DataFrame for mapping validation.

    Applies the same role-based logic as _preview_excel / _preview_csv so that
    the mapping validates against the SAME data that the user sees in the preview.
    """
    ext = path.suffix.lower()

    if ext in _CSV_EXTENSIONS:
        return CsvParser().parse(path).dataframe

    if ext in _EXCEL_EXTENSIONS:
        excel_parser = ExcelParser()
        all_sheets = excel_parser.list_sheets(path) or []
        target = _default_sheet_for_role(source_file.role, all_sheets) or ""
        df = excel_parser.parse(path, sheet_name=target).dataframe

        role = source_file.role
        locator = BlockLocator()

        if role == "amazon_report":
            try:
                return locator.locate_errors_block(df).dataframe
            except BlockNotFoundError:
                return df

        if role == "occ_top" and "plantilla" in target.lower():
            return locator.parse_plantilla(df).dataframe

        return df

    return pd.DataFrame()
