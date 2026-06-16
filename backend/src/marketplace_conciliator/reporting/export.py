"""Export builder — pure domain logic (T-5.3).

Generates xlsx workbooks and csv archives that replicate the three report views.
No I/O or framework imports (ADR-001 hexagonal boundary).
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from typing import Any

import openpyxl

SHEET_FAMILIES = "Errores por familia"
SHEET_SKU_DETAIL = "Detalle SKU"
SHEET_CATALOG_HEALTH = "Salud de catálogo"

_FAMILIES_HEADERS = (
    "Código familia",
    "Familia",
    "Código error",
    "Mensaje",
    "Recuento errores",
    "SKUs únicos (familia)",
)
_SKU_DETAIL_HEADERS = (
    "SKU",
    "SKU normalizado",
    "Código error",
    "Categoría",
    "Mensaje",
    "Campo afectado",
)
_CATALOG_HEADERS = (
    "SKU",
    "SKU normalizado",
    "Estado sync",
    "Stock feed",
    "Stock OCC",
    "Conflicto stock",
    "Stock disponible",
    "En OCC",
    "En feed",
    "En reporte Amazon",
)


@dataclass(frozen=True, slots=True)
class FamilyExportRow:
    """One family/code aggregation row for Vista 1 export."""

    family_code: str
    family_name: str
    error_code: str
    message: str
    error_count: int
    unique_skus_in_family: int


@dataclass(frozen=True, slots=True)
class SkuDetailExportRow:
    """One SKU error row for Vista 2 export."""

    sku_raw: str
    sku_norm: str
    error_code: str
    error_category: str
    error_message: str
    affected_field: str | None


@dataclass(frozen=True, slots=True)
class CatalogHealthExportRow:
    """One catalog-health row for Vista 3 export."""

    sku_raw: str
    sku_norm: str
    sync_status: str
    feed_stock: int | None
    occ_stock: int | None
    stock_conflict: bool
    stock_disponible: bool
    in_occ: bool
    in_feed: bool
    in_amazon_report: bool


@dataclass(frozen=True, slots=True)
class ExportPayload:
    """Structured data for all three export views."""

    families: list[FamilyExportRow]
    sku_details: list[SkuDetailExportRow]
    catalog_health: list[CatalogHealthExportRow]


def _write_families_sheet(ws: Any, rows: list[FamilyExportRow]) -> None:  # noqa: ANN401
    ws.append(list(_FAMILIES_HEADERS))
    for row in rows:
        ws.append(
            [
                row.family_code,
                row.family_name,
                row.error_code,
                row.message,
                row.error_count,
                row.unique_skus_in_family,
            ],
        )


def _write_sku_detail_sheet(ws: Any, rows: list[SkuDetailExportRow]) -> None:  # noqa: ANN401
    ws.append(list(_SKU_DETAIL_HEADERS))
    for row in rows:
        ws.append(
            [
                row.sku_raw,
                row.sku_norm,
                row.error_code,
                row.error_category,
                row.error_message,
                row.affected_field or "",
            ],
        )


def _write_catalog_sheet(ws: Any, rows: list[CatalogHealthExportRow]) -> None:  # noqa: ANN401
    ws.append(list(_CATALOG_HEADERS))
    for row in rows:
        ws.append(
            [
                row.sku_raw,
                row.sku_norm,
                row.sync_status,
                row.feed_stock,
                row.occ_stock,
                row.stock_conflict,
                row.stock_disponible,
                row.in_occ,
                row.in_feed,
                row.in_amazon_report,
            ],
        )


def _rows_to_csv(headers: tuple[str, ...], rows: list[list[object]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def build_xlsx(payload: ExportPayload) -> bytes:
    """Build an xlsx workbook with three sheets mirroring the report tabs."""
    wb = openpyxl.Workbook()
    families_ws = wb.active
    if families_ws is None:
        msg = "openpyxl Workbook has no active sheet"
        raise RuntimeError(msg)
    families_ws.title = SHEET_FAMILIES
    _write_families_sheet(families_ws, payload.families)

    sku_ws = wb.create_sheet(SHEET_SKU_DETAIL)
    _write_sku_detail_sheet(sku_ws, payload.sku_details)

    catalog_ws = wb.create_sheet(SHEET_CATALOG_HEALTH)
    _write_catalog_sheet(catalog_ws, payload.catalog_health)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_csv_archive(payload: ExportPayload) -> bytes:
    """Build a zip archive containing one csv file per report view."""
    families_csv = _rows_to_csv(
        _FAMILIES_HEADERS,
        [
            [
                row.family_code,
                row.family_name,
                row.error_code,
                row.message,
                row.error_count,
                row.unique_skus_in_family,
            ]
            for row in payload.families
        ],
    )
    sku_csv = _rows_to_csv(
        _SKU_DETAIL_HEADERS,
        [
            [
                row.sku_raw,
                row.sku_norm,
                row.error_code,
                row.error_category,
                row.error_message,
                row.affected_field or "",
            ]
            for row in payload.sku_details
        ],
    )
    catalog_csv = _rows_to_csv(
        _CATALOG_HEADERS,
        [
            [
                row.sku_raw,
                row.sku_norm,
                row.sync_status,
                row.feed_stock,
                row.occ_stock,
                row.stock_conflict,
                row.stock_disponible,
                row.in_occ,
                row.in_feed,
                row.in_amazon_report,
            ]
            for row in payload.catalog_health
        ],
    )

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Errores_por_familia.csv", families_csv)
        zf.writestr("Detalle_SKU.csv", sku_csv)
        zf.writestr("Salud_de_catalogo.csv", catalog_csv)
    return archive.getvalue()
