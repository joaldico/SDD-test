"""T-5.3 — BDD CA-05: Informe agregado por familias de error (gate M5).

Suite pytest-bdd ejecutando literalmente los 3 escenarios Gherkin del spec 2.11
(CA-05) sobre la API real (TestClient + SQLite in-memory).

Escenarios cubiertos:
  1. La familia de marca agrega todos sus códigos (Vista 1 drill-down)
  2. Un código desconocido nunca desaparece del informe (SIN_CLASIFICAR + Vista 2)
  3. La exportación replica la estructura de pestañas (xlsx con 3 hojas)

DoD: CA-05 100% verde; los Gherkin del spec se ejecutan literalmente sin reescritura.
"""

from __future__ import annotations

import io
import shutil
from collections.abc import Generator
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from pytest_bdd import given, parsers, scenarios, then, when
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from marketplace_conciliator.ingestion.router import (
    get_db,
    get_db_factory,
    get_staging_dir,
    get_task_runner,
)
from marketplace_conciliator.main import create_app
from marketplace_conciliator.platform.deps import DUMMY_USER
from marketplace_conciliator.reporting.export import (
    SHEET_CATALOG_HEALTH,
    SHEET_FAMILIES,
    SHEET_SKU_DETAIL,
)

scenarios("../features/ca_05_error_families.feature")

_FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"
_CSV_FIXTURE = _FIXTURES_DIR / "wavemarket_fullstock_anonymized.csv"
_XLSX_FIXTURE = _FIXTURES_DIR / "occ_top_sales_anonymized.xlsx"
_XLSM_FIXTURE = _FIXTURES_DIR / "amazon_processing_summary_anonymized.xlsm"

_STAGING = Path(__file__).parent.parent / ".staging_bdd_ca05"
_STAGING.mkdir(parents=True, exist_ok=True)

_engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)


@event.listens_for(_engine, "connect")
def _set_pragma(dbapi_conn: Any, _: Any) -> None:  # noqa: ANN401
    cursor = dbapi_conn.cursor()
    cursor.execute("PRAGMA foreign_keys=ON")
    cursor.close()


def _seed_taxonomy(conn: Any) -> None:  # noqa: ANN401
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS error_families (
            code TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            description TEXT,
            sort_order INTEGER NOT NULL DEFAULT 99
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS error_codes (
            code TEXT PRIMARY KEY,
            family_code TEXT NOT NULL DEFAULT 'SIN_CLASIFICAR'
                REFERENCES error_families(code),
            default_category TEXT,
            canonical_message TEXT,
            first_seen_at DATETIME
        )
    """))
    conn.execute(text("""
        INSERT OR IGNORE INTO error_families (code, display_name, sort_order)
        VALUES
            ('AUTORIZACION_MARCA', 'Autorización de marca', 1),
            ('SIN_CLASIFICAR', 'Sin clasificar', 99)
    """))
    conn.execute(text("""
        INSERT OR IGNORE INTO error_codes (code, family_code, canonical_message)
        VALUES
            ('18299', 'AUTORIZACION_MARCA', 'Marca no autorizada'),
            ('18749', 'AUTORIZACION_MARCA', 'Uso indebido de marca')
    """))


with _engine.begin() as _conn:
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            role TEXT NOT NULL,
            hashed_password TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """))
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS reconciliation_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            marketplace TEXT NOT NULL DEFAULT 'amazon_es',
            status TEXT NOT NULL,
            phase TEXT,
            failure_reason TEXT,
            summary_metrics TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            completed_at DATETIME
        )
    """))
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS source_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL REFERENCES reconciliation_runs(id),
            role TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            sha256 TEXT NOT NULL,
            detected_encoding TEXT,
            detected_delimiter TEXT,
            sheet_name TEXT,
            data_start_row INTEGER,
            header_fingerprint TEXT,
            total_rows INTEGER NOT NULL DEFAULT 0,
            discarded_rows INTEGER NOT NULL DEFAULT 0,
            uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (run_id, role)
        )
    """))
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS column_mappings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file_id INTEGER NOT NULL REFERENCES source_files(id),
            logical_field TEXT NOT NULL,
            source_column_name TEXT NOT NULL,
            source_column_index INTEGER NOT NULL,
            was_suggested INTEGER NOT NULL DEFAULT 0,
            confirmed_by INTEGER NOT NULL REFERENCES users(id),
            confirmed_at DATETIME NOT NULL,
            UNIQUE (source_file_id, logical_field)
        )
    """))
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS run_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL REFERENCES reconciliation_runs(id),
            sku_norm TEXT NOT NULL,
            sku_raw TEXT NOT NULL,
            in_occ INTEGER NOT NULL DEFAULT 0,
            in_feed INTEGER NOT NULL DEFAULT 0,
            in_amazon_report INTEGER NOT NULL DEFAULT 0,
            sync_status TEXT NOT NULL DEFAULT 'NOT_SENT',
            feed_stock INTEGER,
            occ_stock INTEGER,
            stock_conflict INTEGER NOT NULL DEFAULT 0,
            submission_status TEXT,
            UNIQUE (run_id, sku_norm)
        )
    """))
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS duplicate_findings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file_id INTEGER NOT NULL REFERENCES source_files(id),
            sku_norm TEXT NOT NULL,
            occurrences INTEGER NOT NULL,
            resolution TEXT NOT NULL,
            discarded_values TEXT NOT NULL
        )
    """))
    _seed_taxonomy(_conn)
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS item_errors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_item_id INTEGER NOT NULL REFERENCES run_items(id),
            error_code TEXT NOT NULL REFERENCES error_codes(code),
            error_category TEXT NOT NULL DEFAULT 'ERROR',
            error_message TEXT NOT NULL,
            affected_field TEXT
        )
    """))
    _conn.execute(
        text(
            "INSERT OR IGNORE INTO users (id, email, role, hashed_password) "
            "VALUES (:id, :email, :role, 'dummy')",
        ),
        {"id": DUMMY_USER.id, "email": DUMMY_USER.email, "role": DUMMY_USER.role},
    )

_SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=_engine)


def _get_test_db() -> Generator[Session, None, None]:
    db = _SessionLocal()
    try:
        yield db
    finally:
        db.close()


class _SyncTaskRunner:
    """Runs the pipeline synchronously in-process (same pattern as CA-02)."""

    def submit(self, run_id: int, work_fn: Any) -> None:  # noqa: ANN401
        work_fn(run_id)

    def active_count(self) -> int:
        return 0

    def shutdown(self, *, wait: bool = True) -> None:  # noqa: ARG002
        pass


_sync_runner = _SyncTaskRunner()


@pytest.fixture(scope="module")
def bdd_client() -> TestClient:
    app = create_app()
    app.dependency_overrides[get_db] = _get_test_db
    app.dependency_overrides[get_staging_dir] = lambda: _STAGING
    app.dependency_overrides[get_task_runner] = lambda: _sync_runner
    app.dependency_overrides[get_db_factory] = lambda: _SessionLocal
    return TestClient(app)


@pytest.fixture
def ctx() -> dict[str, Any]:
    return {}


def _make_csv(rows: list[dict[str, str]], *, header: list[str] | None = None) -> bytes:
    cols = header or list(rows[0].keys())
    lines = [",".join(cols)]
    lines.extend(",".join(str(row.get(c, "")) for c in cols) for row in rows)
    return "\n".join(lines).encode("utf-8")


def _make_occ_xlsx(rows: list[dict[str, str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(["Name", "SKU", "Supplier", None, "stock occ"])
    for row in rows:
        ws.append([
            row.get("Name", ""),
            row.get("SKU", ""),
            row.get("Supplier", ""),
            None,
            row.get("stock occ", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_amazon_report_xlsx(error_rows: list[dict[str, str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen de procesamiento"
    ws.append([""])
    ws.append(["Errores y advertencias por SKU"])
    ws.append([
        "#", "Código de error", "Categoría de error",
        "Mensaje de error", "Campo afectado", "SKU",
    ])
    for i, row in enumerate(error_rows, start=1):
        ws.append([
            str(i),
            row.get("Código de error", ""),
            row.get("Categoría de error", "ERROR"),
            row.get("Mensaje de error", ""),
            row.get("Campo afectado", ""),
            row.get("SKU", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_file(  # noqa: PLR0913
    client: TestClient,
    run_id: int,
    role: str,
    filename: str,
    data: bytes,
    content_type: str = "application/octet-stream",
) -> int:
    resp = client.post(
        f"/api/v1/runs/{run_id}/files",
        data={"role": role},
        files={"file": (filename, io.BytesIO(data), content_type)},
    )
    assert resp.status_code == 201, f"Upload failed ({role}): {resp.text}"
    return int(resp.json()["id"])


def _confirm_mapping(
    client: TestClient,
    run_id: int,
    file_id: int,
    mappings: list[dict[str, Any]],
) -> None:
    resp = client.put(
        f"/api/v1/runs/{run_id}/files/{file_id}/mapping",
        json={"mappings": mappings},
    )
    assert resp.status_code == 200, f"Mapping confirmation failed: {resp.text}"


def _confirm_sku_by_preview(client: TestClient, run_id: int, file_id: int) -> None:
    resp = client.get(f"/api/v1/runs/{run_id}/files/{file_id}/preview")
    assert resp.status_code == 200, f"Preview failed: {resp.text}"
    suggestions = resp.json()["suggestions"]
    mappings = [
        {"logical_field": field, "column_index": info["column_index"], "was_suggested": True}
        for field, info in suggestions.items()
    ]
    if not mappings:
        mappings = [{"logical_field": "sku", "column_index": 0, "was_suggested": False}]
    _confirm_mapping(client, run_id, file_id, mappings)


def _create_run(client: TestClient) -> int:
    resp = client.post("/api/v1/runs", json={})
    assert resp.status_code == 201
    return int(resp.json()["id"])


def _insert_completed_run_with_brand_errors(
    db: Session,
    *,
    count_18299: int,
    count_18749: int,
) -> int:
    db.execute(
        text("""
            INSERT INTO reconciliation_runs
                (user_id, marketplace, status, completed_at)
            VALUES
                (1, 'amazon_es', 'completed', '2026-06-16 12:00:00')
        """),
    )
    db.commit()
    run_id = int(db.execute(text("SELECT last_insert_rowid()")).scalar_one())

    db.execute(
        text("""
            INSERT INTO run_items
                (run_id, sku_norm, sku_raw, in_occ, in_feed, in_amazon_report,
                 sync_status, feed_stock, occ_stock, stock_conflict)
            VALUES
                (:run_id, 'SKU-18299', 'SKU-18299', 0, 1, 1, 'SENT_WITH_ERROR', 5, NULL, 0),
                (:run_id, 'SKU-18749', 'SKU-18749', 0, 1, 1, 'SENT_WITH_ERROR', 3, NULL, 0)
        """),
        {"run_id": run_id},
    )
    item_a = int(
        db.execute(
            text("SELECT id FROM run_items WHERE run_id = :run_id AND sku_norm = 'SKU-18299'"),
            {"run_id": run_id},
        ).scalar_one(),
    )
    item_b = int(
        db.execute(
            text("SELECT id FROM run_items WHERE run_id = :run_id AND sku_norm = 'SKU-18749'"),
            {"run_id": run_id},
        ).scalar_one(),
    )

    errors: list[dict[str, Any]] = []
    for idx in range(count_18299):
        errors.append({
            "item_id": item_a,
            "code": "18299",
            "category": "ERROR",
            "message": f"Marca no autorizada #{idx}",
            "field": "brand",
        })
    for idx in range(count_18749):
        errors.append({
            "item_id": item_b,
            "code": "18749",
            "category": "ERROR",
            "message": f"Uso indebido de marca #{idx}",
            "field": "title",
        })

    db.connection().execute(
        text("""
            INSERT INTO item_errors
                (run_item_id, error_code, error_category, error_message, affected_field)
            VALUES
                (:item_id, :code, :category, :message, :field)
        """),
        errors,
    )
    db.commit()
    return run_id


def _setup_full_fixture_run(client: TestClient) -> int:
    run_id = _create_run(client)

    wm_id = _upload_file(
        client, run_id, "wm_feed", _CSV_FIXTURE.name,
        _CSV_FIXTURE.read_bytes(), "text/csv",
    )
    prev = client.get(f"/api/v1/runs/{run_id}/files/{wm_id}/preview").json()
    wm_mappings: list[dict[str, Any]] = [
        {
            "logical_field": "sku",
            "column_index": prev["suggestions"]["sku"]["column_index"],
            "was_suggested": True,
        },
    ]
    stock_idx = prev["suggestions"].get("stock", {}).get("column_index")
    if stock_idx is not None:
        wm_mappings.append(
            {"logical_field": "stock", "column_index": stock_idx, "was_suggested": True},
        )
    _confirm_mapping(client, run_id, wm_id, wm_mappings)

    occ_id = _upload_file(
        client, run_id, "occ_top", _XLSX_FIXTURE.name,
        _XLSX_FIXTURE.read_bytes(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    prev = client.get(f"/api/v1/runs/{run_id}/files/{occ_id}/preview").json()
    occ_mappings: list[dict[str, Any]] = [
        {
            "logical_field": "sku",
            "column_index": prev["suggestions"]["sku"]["column_index"],
            "was_suggested": True,
        },
    ]
    occ_stock_idx = prev["suggestions"].get("stock", {}).get("column_index")
    if occ_stock_idx is not None:
        occ_mappings.append(
            {"logical_field": "stock", "column_index": occ_stock_idx, "was_suggested": True},
        )
    _confirm_mapping(client, run_id, occ_id, occ_mappings)

    amz_id = _upload_file(
        client, run_id, "amazon_report", _XLSM_FIXTURE.name,
        _XLSM_FIXTURE.read_bytes(),
        "application/vnd.ms-excel.sheet.macroenabled.12",
    )
    _confirm_sku_by_preview(client, run_id, amz_id)
    return run_id


# =============================================================================
# Escenario 1: La familia de marca agrega todos sus códigos
# =============================================================================


@given(parsers.parse('que la conciliación produjo {count:d} errores con código "{code}"'))
def step_s1_produced_errors(ctx: dict[str, Any], count: int, code: str) -> None:
    ctx.setdefault("error_counts", {})[code] = count


@given(parsers.parse('{count:d} errores con código "{code}"'))
def step_s1_additional_errors(ctx: dict[str, Any], count: int, code: str) -> None:
    ctx.setdefault("error_counts", {})[code] = count


@given(parsers.parse('ambos códigos pertenecen a la familia "{family_code}"'))
def step_s1_family_codes(
    bdd_client: TestClient,  # noqa: ARG001
    ctx: dict[str, Any],
    family_code: str,
) -> None:
    counts = ctx.get("error_counts", {})
    assert "18299" in counts and "18749" in counts

    db = _SessionLocal()
    try:
        run_id = _insert_completed_run_with_brand_errors(
            db,
            count_18299=int(counts["18299"]),
            count_18749=int(counts["18749"]),
        )
    finally:
        db.close()

    ctx["run_id"] = run_id
    ctx["family_code"] = family_code


@when('abro la Vista 1 "Errores por familia"')
def step_s1_open_vista1(bdd_client: TestClient, ctx: dict[str, Any]) -> None:
    resp = bdd_client.get(f"/api/v1/runs/{ctx['run_id']}/report/families")
    assert resp.status_code == 200, resp.text
    ctx["families_report"] = resp.json()


@then(parsers.parse('veo la familia "{display_name}" con el total de SKUs únicos afectados'))
def step_s1_see_family(ctx: dict[str, Any], display_name: str) -> None:
    report = ctx["families_report"]
    family = next(
        (f for f in report["families"] if f["display_name"] == display_name),
        None,
    )
    assert family is not None, f"Familia '{display_name}' no encontrada"
    assert family["unique_skus"] == 2
    ctx["selected_family"] = family


@then(
    parsers.parse(
        'al desplegarla veo el desglose por código: "{code_a}" y "{code_b}" con sus recuentos',
    ),
)
def step_s1_code_breakdown(ctx: dict[str, Any], code_a: str, code_b: str) -> None:
    family = ctx["selected_family"]
    codes = {c["code"]: c["count"] for c in family["codes"]}
    expected = ctx["error_counts"]
    assert code_a in codes and code_b in codes
    assert codes[code_a] == expected[code_a]
    assert codes[code_b] == expected[code_b]
    ctx["selected_code"] = code_a


@then("al seleccionar un código veo la lista de SKUs afectados con la descripción del error")
def step_s1_sku_list_for_code(ctx: dict[str, Any]) -> None:
    code = ctx["selected_code"]
    db = _SessionLocal()
    try:
        rows = db.execute(
            text("""
                SELECT ri.sku_norm, ie.error_message
                FROM item_errors ie
                JOIN run_items ri ON ri.id = ie.run_item_id
                WHERE ri.run_id = :run_id AND ie.error_code = :code
                ORDER BY ri.sku_norm ASC
                LIMIT 5
            """),
            {"run_id": ctx["run_id"], "code": code},
        ).fetchall()
    finally:
        db.close()

    assert rows, f"No hay SKUs para el código {code}"
    for sku, message in rows:
        assert sku
        assert message


# =============================================================================
# Escenario 2: Un código desconocido nunca desaparece del informe
# =============================================================================


@given(
    parsers.parse(
        'que el reporte de Amazon contiene el código "{code}" que no existe en el catálogo',
    ),
)
def step_s2_unknown_code_in_report(bdd_client: TestClient, ctx: dict[str, Any], code: str) -> None:
    run_id = _create_run(bdd_client)
    ctx["run_id"] = run_id
    ctx["unknown_code"] = code
    ctx["unknown_sku"] = "SKU-UNKNOWN-999"

    feed_csv = _make_csv([
        {"sku": ctx["unknown_sku"], "stock": "4", "site": "ES", "condition": "new"},
    ])
    feed_id = _upload_file(bdd_client, run_id, "wm_feed", "feed.csv", feed_csv, "text/csv")
    _confirm_mapping(bdd_client, run_id, feed_id, [
        {"logical_field": "sku", "column_index": 0, "was_suggested": True},
        {"logical_field": "stock", "column_index": 1, "was_suggested": True},
    ])

    occ_xlsx = _make_occ_xlsx([{"Name": "Neutral", "SKU": "NEUTRAL_OCC", "Supplier": "S1"}])
    occ_id = _upload_file(
        bdd_client, run_id, "occ_top", "occ.xlsx", occ_xlsx,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    _confirm_mapping(bdd_client, run_id, occ_id, [
        {"logical_field": "sku", "column_index": 1, "was_suggested": True},
    ])

    amazon_xlsx = _make_amazon_report_xlsx([{
        "Código de error": code,
        "Categoría de error": "ERROR",
        "Mensaje de error": "Error desconocido de prueba",
        "Campo afectado": "title",
        "SKU": ctx["unknown_sku"],
    }])
    amazon_id = _upload_file(
        bdd_client, run_id, "amazon_report", "report.xlsx", amazon_xlsx,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    _confirm_sku_by_preview(bdd_client, run_id, amazon_id)


@when("finaliza la conciliación", target_fixture="process_resp")
def step_s2_finaliza_conciliacion(
    bdd_client: TestClient,
    ctx: dict[str, Any],
) -> dict[str, Any]:
    resp = bdd_client.post(f"/api/v1/runs/{ctx['run_id']}/process")
    assert resp.status_code == 202, resp.text
    return resp.json()


@then(
    parsers.parse(
        'el código "{code}" se registra en el catálogo asignado a la familia "{family_code}"',
    ),
)
def step_s2_code_registered(ctx: dict[str, Any], code: str, family_code: str) -> None:
    db = _SessionLocal()
    try:
        row = db.execute(
            text("SELECT family_code FROM error_codes WHERE code = :code"),
            {"code": code},
        ).fetchone()
    finally:
        db.close()

    assert row is not None, f"Código {code} no registrado en error_codes"
    assert row[0] == family_code


@then(parsers.parse('la Vista 1 muestra la familia "{display_name}" con un aviso visible'))
def step_s2_vista1_warning(bdd_client: TestClient, ctx: dict[str, Any], display_name: str) -> None:
    resp = bdd_client.get(f"/api/v1/runs/{ctx['run_id']}/report/families")
    assert resp.status_code == 200
    data = resp.json()
    assert data["sin_clasificar_warning"] is True
    family = next((f for f in data["families"] if f["display_name"] == display_name), None)
    assert family is not None
    assert family["total_errors"] >= 1


@then(parsers.parse('los SKUs afectados por "{code}" conservan su detalle completo en la Vista 2'))
def step_s2_vista2_detail(ctx: dict[str, Any], code: str) -> None:
    db = _SessionLocal()
    try:
        rows = db.execute(
            text("""
                SELECT ie.error_code, ie.error_category, ie.error_message, ie.affected_field
                FROM item_errors ie
                JOIN run_items ri ON ri.id = ie.run_item_id
                WHERE ri.run_id = :run_id AND ie.error_code = :code
            """),
            {"run_id": ctx["run_id"], "code": code},
        ).fetchall()
    finally:
        db.close()

    assert len(rows) >= 1
    for error_code, category, message, field in rows:
        assert error_code == code
        assert category in ("ERROR", "ADVERTENCIA")
        assert message
        assert field is not None


# =============================================================================
# Escenario 3: La exportación replica la estructura de pestañas
# =============================================================================


@given("que la conciliación está completada")
def step_s3_completed_reconciliation(bdd_client: TestClient, ctx: dict[str, Any]) -> None:
    run_id = _setup_full_fixture_run(bdd_client)
    resp = bdd_client.post(f"/api/v1/runs/{run_id}/process")
    assert resp.status_code == 202, resp.text

    status = bdd_client.get(f"/api/v1/runs/{run_id}/status").json()
    assert status["status"] == "completed", status

    db = _SessionLocal()
    try:
        ctx["expected_families_rows"] = db.execute(
            text("""
                SELECT COUNT(*)
                FROM (
                    SELECT ef.code, ec.code
                    FROM item_errors ie
                    JOIN run_items ri ON ri.id = ie.run_item_id
                    JOIN error_codes ec ON ec.code = ie.error_code
                    JOIN error_families ef ON ef.code = ec.family_code
                    WHERE ri.run_id = :run_id
                    GROUP BY ef.code, ec.code
                )
            """),
            {"run_id": run_id},
        ).scalar_one()
        ctx["expected_sku_rows"] = db.execute(
            text("""
                SELECT COUNT(*)
                FROM item_errors ie
                JOIN run_items ri ON ri.id = ie.run_item_id
                WHERE ri.run_id = :run_id
            """),
            {"run_id": run_id},
        ).scalar_one()
        ctx["expected_catalog_rows"] = db.execute(
            text("SELECT COUNT(*) FROM run_items WHERE run_id = :run_id"),
            {"run_id": run_id},
        ).scalar_one()
    finally:
        db.close()

    ctx["run_id"] = run_id


@when("exporto el informe a formato xlsx")
def step_s3_export_xlsx(bdd_client: TestClient, ctx: dict[str, Any]) -> None:
    resp = bdd_client.get(
        f"/api/v1/runs/{ctx['run_id']}/export",
        params={"format": "xlsx"},
    )
    assert resp.status_code == 200, resp.text
    ctx["export_workbook"] = openpyxl.load_workbook(io.BytesIO(resp.content))


@then("el libro contiene una pestaña con la agregación por familia y código")
def step_s3_families_sheet(ctx: dict[str, Any]) -> None:
    wb = ctx["export_workbook"]
    assert SHEET_FAMILIES in wb.sheetnames
    ws = wb[SHEET_FAMILIES]
    headers = [cell.value for cell in ws[1]]
    assert "Código familia" in headers
    assert "Código error" in headers
    data_rows = ws.max_row - 1
    assert data_rows == ctx["expected_families_rows"]


@then("otra pestaña con el detalle SKU, código de error y descripción")
def step_s3_sku_detail_sheet(ctx: dict[str, Any]) -> None:
    wb = ctx["export_workbook"]
    assert SHEET_SKU_DETAIL in wb.sheetnames
    ws = wb[SHEET_SKU_DETAIL]
    headers = [cell.value for cell in ws[1]]
    assert "SKU" in headers
    assert "Código error" in headers
    assert "Mensaje" in headers
    assert ws.max_row - 1 == ctx["expected_sku_rows"]


@then("otra pestaña con la salud de catálogo")
def step_s3_catalog_sheet(ctx: dict[str, Any]) -> None:
    wb = ctx["export_workbook"]
    assert SHEET_CATALOG_HEALTH in wb.sheetnames
    ws = wb[SHEET_CATALOG_HEALTH]
    headers = [cell.value for cell in ws[1]]
    assert "Estado sync" in headers
    assert "Stock feed" in headers
    assert ws.max_row - 1 == ctx["expected_catalog_rows"]


def teardown_module() -> None:
    if _STAGING.exists():
        shutil.rmtree(_STAGING, ignore_errors=True)
