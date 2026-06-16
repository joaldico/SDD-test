from __future__ import annotations

import csv
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

FIXTURES_DIR = Path(__file__).parent / "fixtures"
EXPECTED_OCC_ROWS = 1_232
EXPECTED_FEED_ROWS = 4_156
EXPECTED_REPORT_ROWS = 8_173
EXPECTED_NBSP_MESSAGES = 463
EXPECTED_S01098S3MRN_ERRORS = 11
EXPECTED_REPORT_SHEETS = [
    "Instrucciones",
    "Resumen de procesamiento",
    "Definiciones de datos",
    "Dropdown Lists",
    "AttributePTDMAP",
    "Plantilla",
    "Conditions List",
    "Valores válidos",
]
EXPECTED_REPORT_HEADER = (
    None,
    "#",
    "Código de error",
    "Categoría de error",
    "Mensaje de error",
    "Campo afectado (celda impactada)",
    "SKU",
)


def test_canonical_fixtures_open_and_match_known_row_counts() -> None:
    occ_path = FIXTURES_DIR / "occ_top_sales_anonymized.xlsx"
    feed_path = FIXTURES_DIR / "wavemarket_fullstock_anonymized.csv"
    report_path = FIXTURES_DIR / "amazon_processing_summary_anonymized.xlsm"

    occ_workbook = load_workbook(occ_path, read_only=True, data_only=True)
    occ_sheet = occ_workbook["Hoja1"]
    occ_rows = list(occ_sheet.iter_rows(min_row=2, values_only=True))

    with feed_path.open(encoding="utf-8", newline="") as feed_file:
        feed_rows = list(csv.DictReader(feed_file))

    report_workbook = load_workbook(report_path, read_only=True, data_only=True, keep_vba=True)
    summary_sheet = report_workbook["Resumen de procesamiento"]
    report_rows = [
        row
        for row in summary_sheet.iter_rows(min_row=572, max_col=7, values_only=True)
        if row[1] is not None
    ]

    assert len(occ_rows) == EXPECTED_OCC_ROWS
    assert len(feed_rows) == EXPECTED_FEED_ROWS
    assert len(report_rows) == EXPECTED_REPORT_ROWS
    assert {"03763BAR", "K2.65"}.issubset({row["sku"] for row in feed_rows})
    assert report_workbook.sheetnames == EXPECTED_REPORT_SHEETS
    assert summary_sheet["B570"].value == "Errores y advertencias por SKU"
    assert (
        next(summary_sheet.iter_rows(min_row=571, max_row=571, max_col=7, values_only=True))
        == EXPECTED_REPORT_HEADER
    )
    assert sum("\xa0" in str(row[4]) for row in report_rows) == EXPECTED_NBSP_MESSAGES
    assert sum(row[6] == "S01098S3MRN" for row in report_rows) == EXPECTED_S01098S3MRN_ERRORS

    plantilla_sheet = report_workbook["Plantilla"]
    assert plantilla_sheet["D4"].value == "SKU"
    assert plantilla_sheet["D5"].value == "contribution_sku#1.value"
    assert plantilla_sheet["D6"].value == "ABC123"

    with ZipFile(report_path) as archive:
        assert "xl/vbaProject.bin" not in archive.namelist()
