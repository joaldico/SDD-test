"""T-5.3 — Unit tests for reporting.export (pure domain logic, TDD)."""

from __future__ import annotations

import io
import zipfile

import openpyxl

from marketplace_conciliator.reporting.export import (
    SHEET_CATALOG_HEALTH,
    SHEET_FAMILIES,
    SHEET_SKU_DETAIL,
    CatalogHealthExportRow,
    ExportPayload,
    FamilyExportRow,
    SkuDetailExportRow,
    build_csv_archive,
    build_xlsx,
)


def _sample_payload() -> ExportPayload:
    return ExportPayload(
        families=[
            FamilyExportRow(
                family_code="AUTORIZACION_MARCA",
                family_name="Autorización de marca",
                error_code="18299",
                message="Marca no autorizada",
                error_count=1786,
                unique_skus_in_family=950,
            ),
            FamilyExportRow(
                family_code="AUTORIZACION_MARCA",
                family_name="Autorización de marca",
                error_code="18749",
                message="Uso indebido de marca",
                error_count=118,
                unique_skus_in_family=950,
            ),
        ],
        sku_details=[
            SkuDetailExportRow(
                sku_raw="SKU-A",
                sku_norm="SKU-A",
                error_code="18299",
                error_category="ERROR",
                error_message="Marca no autorizada en SKU-A",
                affected_field="brand",
            ),
            SkuDetailExportRow(
                sku_raw="SKU-B",
                sku_norm="SKU-B",
                error_code="18749",
                error_category="ERROR",
                error_message="Uso indebido de marca",
                affected_field="title",
            ),
        ],
        catalog_health=[
            CatalogHealthExportRow(
                sku_raw="AAA111",
                sku_norm="AAA111",
                sync_status="DESYNC_FEED_ONLY",
                feed_stock=25,
                occ_stock=None,
                stock_conflict=False,
                stock_disponible=True,
                in_occ=False,
                in_feed=True,
                in_amazon_report=False,
            ),
        ],
    )


class TestBuildXlsx:
    def test_workbook_has_three_sheets_with_expected_names(self) -> None:
        payload = _sample_payload()
        content = build_xlsx(payload)

        wb = openpyxl.load_workbook(io.BytesIO(content))
        assert wb.sheetnames == [SHEET_FAMILIES, SHEET_SKU_DETAIL, SHEET_CATALOG_HEALTH]

    def test_families_sheet_contains_family_and_code_rows(self) -> None:
        payload = _sample_payload()
        wb = openpyxl.load_workbook(io.BytesIO(build_xlsx(payload)))
        ws = wb[SHEET_FAMILIES]

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        assert rows[0][0] == "AUTORIZACION_MARCA"
        assert rows[0][2] == "18299"
        assert rows[0][4] == 1786
        assert rows[1][2] == "18749"
        assert rows[1][4] == 118

    def test_sku_detail_sheet_has_one_row_per_error(self) -> None:
        payload = _sample_payload()
        wb = openpyxl.load_workbook(io.BytesIO(build_xlsx(payload)))
        ws = wb[SHEET_SKU_DETAIL]

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        assert rows[0][0] == "SKU-A"
        assert rows[0][2] == "18299"
        assert rows[1][0] == "SKU-B"

    def test_catalog_health_sheet_replicates_view_columns(self) -> None:
        payload = _sample_payload()
        wb = openpyxl.load_workbook(io.BytesIO(build_xlsx(payload)))
        ws = wb[SHEET_CATALOG_HEALTH]

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][0] == "AAA111"
        assert rows[0][2] == "DESYNC_FEED_ONLY"
        assert rows[0][3] == 25


class TestBuildCsvArchive:
    def test_archive_contains_three_csv_files(self) -> None:
        payload = _sample_payload()
        archive = build_csv_archive(payload)

        with zipfile.ZipFile(io.BytesIO(archive)) as zf:
            names = set(zf.namelist())
            assert names == {
                "Errores_por_familia.csv",
                "Detalle_SKU.csv",
                "Salud_de_catalogo.csv",
            }

    def test_csv_files_have_header_plus_data_rows(self) -> None:
        payload = _sample_payload()
        archive = build_csv_archive(payload)

        with zipfile.ZipFile(io.BytesIO(archive)) as zf:
            families = zf.read("Errores_por_familia.csv").decode("utf-8-sig")
            sku_detail = zf.read("Detalle_SKU.csv").decode("utf-8-sig")
            catalog = zf.read("Salud_de_catalogo.csv").decode("utf-8-sig")

        assert len(families.strip().splitlines()) == 3  # header + 2 rows
        assert len(sku_detail.strip().splitlines()) == 3
        assert len(catalog.strip().splitlines()) == 2
