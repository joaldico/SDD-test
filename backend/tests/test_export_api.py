"""T-5.3 — Integration tests for GET /api/v1/runs/{id}/export (TDD)."""

from __future__ import annotations

import io
import json
import zipfile
from collections.abc import Generator
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from marketplace_conciliator.main import create_app
from marketplace_conciliator.reporting.export import (
    SHEET_CATALOG_HEALTH,
    SHEET_FAMILIES,
    SHEET_SKU_DETAIL,
)
from marketplace_conciliator.reporting.router import get_db

_engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)


@event.listens_for(_engine, "connect")
def _set_pragma(dbapi_conn: Any, _: Any) -> None:  # noqa: ANN401
    dbapi_conn.cursor().execute("PRAGMA foreign_keys=ON")


def _create_schema(conn: Any) -> None:  # noqa: ANN401
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            role TEXT NOT NULL,
            hashed_password TEXT NOT NULL
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS reconciliation_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            marketplace TEXT NOT NULL DEFAULT 'amazon_es',
            status TEXT NOT NULL,
            phase TEXT,
            failure_reason TEXT,
            summary_metrics TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            completed_at DATETIME
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS error_families (
            code TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            description TEXT,
            sort_order INTEGER NOT NULL
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS error_codes (
            code TEXT PRIMARY KEY,
            family_code TEXT NOT NULL REFERENCES error_families(code),
            default_category TEXT,
            canonical_message TEXT,
            first_seen_at DATETIME
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS run_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL REFERENCES reconciliation_runs(id),
            sku_norm TEXT NOT NULL,
            sku_raw TEXT NOT NULL,
            in_occ INTEGER NOT NULL,
            in_feed INTEGER NOT NULL,
            in_amazon_report INTEGER NOT NULL,
            sync_status TEXT NOT NULL,
            feed_stock INTEGER,
            occ_stock INTEGER,
            stock_conflict INTEGER NOT NULL DEFAULT 0,
            submission_status TEXT,
            UNIQUE(run_id, sku_norm)
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS item_errors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_item_id INTEGER NOT NULL REFERENCES run_items(id),
            error_code TEXT NOT NULL REFERENCES error_codes(code),
            error_category TEXT NOT NULL,
            error_message TEXT NOT NULL,
            affected_field TEXT
        )
    """))
    conn.execute(text("""
        INSERT OR IGNORE INTO users (id, email, role, hashed_password)
        VALUES (1, 'admin@test.local', 'admin', 'dummy')
    """))
    conn.execute(text("""
        INSERT OR IGNORE INTO error_families (code, display_name, sort_order)
        VALUES
            ('AUTORIZACION_MARCA', 'Autorización de marca', 1),
            ('SIN_CLASIFICAR', 'Sin clasificar', 99)
    """))
    conn.execute(text("""
        INSERT OR IGNORE INTO error_codes (code, family_code, canonical_message)
        VALUES
            ('18299', 'AUTORIZACION_MARCA', 'Marca no autorizada'),
            ('18749', 'AUTORIZACION_MARCA', 'Uso indebido de marca')
    """))


with _engine.begin() as conn:
    _create_schema(conn)

_SessionLocal = sessionmaker(bind=_engine, autocommit=False, autoflush=False)


def _get_test_db() -> Generator[Session, None, None]:
    db = _SessionLocal()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture
def client() -> Generator[TestClient, None, None]:
    app = create_app()
    app.dependency_overrides[get_db] = _get_test_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def _insert_completed_run(db: Session, *, summary: dict[str, int] | None = None) -> int:
    metrics = summary or {
        "total_skus": 3,
        "total_errors": 1904,
        "sent_with_error": 2,
        "sent_ok": 0,
        "not_sent": 1,
        "desync_feed_only": 0,
        "desync_amazon_only": 0,
    }
    db.execute(
        text("""
            INSERT INTO reconciliation_runs
                (user_id, marketplace, status, summary_metrics, completed_at)
            VALUES
                (1, 'amazon_es', 'completed', :metrics, '2026-06-16 12:00:00')
        """),
        {"metrics": json.dumps(metrics)},
    )
    db.commit()
    return int(db.execute(text("SELECT last_insert_rowid()")).scalar_one())


def _seed_brand_export_data(db: Session, run_id: int) -> None:
    db.execute(
        text("""
            INSERT INTO run_items
                (run_id, sku_norm, sku_raw, in_occ, in_feed, in_amazon_report,
                 sync_status, feed_stock, occ_stock, stock_conflict)
            VALUES
                (:run_id, 'SKU-A', 'SKU-A', 0, 1, 1, 'SENT_WITH_ERROR', 5, NULL, 0),
                (:run_id, 'SKU-B', 'SKU-B', 0, 1, 1, 'SENT_WITH_ERROR', 3, NULL, 0),
                (:run_id, 'SKU-C', 'SKU-C', 1, 0, 0, 'NOT_SENT', NULL, 2, 0)
        """),
        {"run_id": run_id},
    )
    item_a = int(
        db.execute(
            text("SELECT id FROM run_items WHERE run_id = :run_id AND sku_norm = 'SKU-A'"),
            {"run_id": run_id},
        ).scalar_one(),
    )
    item_b = int(
        db.execute(
            text("SELECT id FROM run_items WHERE run_id = :run_id AND sku_norm = 'SKU-B'"),
            {"run_id": run_id},
        ).scalar_one(),
    )

    errors: list[tuple[int, str, str, str, str | None]] = []
    for idx in range(1786):
        errors.append((item_a, "18299", "ERROR", f"Mensaje 18299 #{idx}", "brand"))
    for idx in range(118):
        errors.append((item_b, "18749", "ERROR", f"Mensaje 18749 #{idx}", "title"))

    params = [
        {
            "item_id": item_id,
            "code": code,
            "category": category,
            "message": message,
            "field": field,
        }
        for item_id, code, category, message, field in errors
    ]
    db.connection().execute(
        text("""
            INSERT INTO item_errors
                (run_item_id, error_code, error_category, error_message, affected_field)
            VALUES
                (:item_id, :code, :category, :message, :field)
        """),
        params,
    )
    db.commit()


class TestExportApi:
    def test_export_xlsx_returns_three_sheets_with_fixture_counts(
        self,
        client: TestClient,
    ) -> None:
        db = _SessionLocal()
        try:
            run_id = _insert_completed_run(db)
            _seed_brand_export_data(db, run_id)
        finally:
            db.close()

        response = client.get(f"/api/v1/runs/{run_id}/export", params={"format": "xlsx"})
        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        assert wb.sheetnames == [SHEET_FAMILIES, SHEET_SKU_DETAIL, SHEET_CATALOG_HEALTH]

        families_rows = list(wb[SHEET_FAMILIES].iter_rows(min_row=2, values_only=True))
        sku_rows = list(wb[SHEET_SKU_DETAIL].iter_rows(min_row=2, values_only=True))
        catalog_rows = list(wb[SHEET_CATALOG_HEALTH].iter_rows(min_row=2, values_only=True))

        assert len(families_rows) == 2
        assert {row[2] for row in families_rows} == {"18299", "18749"}
        counts = {row[2]: row[4] for row in families_rows}
        assert counts["18299"] == 1786
        assert counts["18749"] == 118
        assert len(sku_rows) == 1904
        assert len(catalog_rows) == 3

    def test_export_csv_returns_zip_with_three_files(self, client: TestClient) -> None:
        db = _SessionLocal()
        try:
            run_id = _insert_completed_run(db)
            _seed_brand_export_data(db, run_id)
        finally:
            db.close()

        response = client.get(f"/api/v1/runs/{run_id}/export", params={"format": "csv"})
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/zip"

        with zipfile.ZipFile(io.BytesIO(response.content)) as zf:
            assert len(zf.namelist()) == 3

    def test_export_processing_run_returns_409(self, client: TestClient) -> None:
        db = _SessionLocal()
        try:
            db.execute(
                text("""
                    INSERT INTO reconciliation_runs (user_id, marketplace, status)
                    VALUES (1, 'amazon_es', 'processing')
                """),
            )
            db.commit()
            run_id = int(db.execute(text("SELECT last_insert_rowid()")).scalar_one())
        finally:
            db.close()

        response = client.get(f"/api/v1/runs/{run_id}/export", params={"format": "xlsx"})
        assert response.status_code == 409

    def test_export_invalid_format_returns_422(self, client: TestClient) -> None:
        db = _SessionLocal()
        try:
            run_id = _insert_completed_run(db)
        finally:
            db.close()

        response = client.get(f"/api/v1/runs/{run_id}/export", params={"format": "pdf"})
        assert response.status_code == 422
